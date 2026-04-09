import pandas as pd
from openpyxl import load_workbook
import json

only_first_name = True

workbook_path = "./members.xlsx"
family_names = ["Perry Family", "Tea Family"]
workbook = load_workbook(workbook_path)

print(workbook.sheetnames)


for family_index, family in enumerate(workbook):
    members = []
    members_id = {}
    df = pd.read_excel(workbook_path, sheet_name=family.title)
    names = df["Name"].tolist()
    parents = df["Angel"].tolist()
    statuses = df["Status"].tolist()

    for i, name in enumerate(names):
        members_id[name] = i

    for i in range(len(names)):
        if pd.isna(parents[i]):
            parents[i] = None
        if pd.isna(statuses[i]):
            statuses[i] = "ACTIVE"
        else:
            statuses[i] = statuses[i].upper()
        members.append(
            {
                "name": names[i].split()[0] if only_first_name else names[i],
                "id": i,
                "parent": members_id[parents[i]] if parents[i] in members_id else None,
                "status": statuses[i],
                "Additional Info": "",
            }
        )
    with open(f"./members_{family_index}.json", "w") as f:
        json.dump(members, f, ensure_ascii=False, indent=4)
